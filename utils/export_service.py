"""XLSX/CSV export using pandas + openpyxl."""
import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .defaults import MOVEMENT_TYPE_LABELS, MONTHS, INCOME_AUTO_ITEMS

TYPE_ORDER = ["ingreso", "costo_venta", "gasto_venta", "gasto_operacional", "pago_deuda"]
EXPENSE_TYPES = ["costo_venta", "gasto_venta", "gasto_operacional", "pago_deuda"]

# Colors
COLORS = {
    "ingreso":           "4CAF7D",
    "costo_venta":       "5D80B5",
    "gasto_venta":       "EF599A",
    "gasto_operacional": "A870B0",
    "pago_deuda":        "62CBE6",
    "header_bg":         "141D2E",
    "total_bg":          "1A2540",
    "grand_total_bg":    "0F1520",
}


def _fmt(val):
    return round(float(val), 2) if val else 0.0


def build_matrix(project: dict, movements: list, year: int) -> dict:
    """
    Build the summary matrix:
    rows = items (by category), columns = months + totals + avg
    Returns dict with rows, type_subtotals, month_keys.
    """
    year_str = str(year)
    month_keys = [m for m, _ in MONTHS]

    filtered = [m for m in movements if m["project_id"] == project["id"]
                and m["date"].startswith(year_str)]

    # Totals per item per month
    item_month = {}  # {item_id: {month: amount}}
    for mv in filtered:
        mid = mv["item_id"]
        month = mv["date"][5:7]
        item_month.setdefault(mid, {})
        item_month[mid][month] = item_month[mid].get(month, 0) + mv["amount"]

    # Saldo mes anterior (carry-forward)
    saldo = {}
    cumulative = 0.0
    for month in month_keys:
        saldo[month] = cumulative
        inc = sum(m["amount"] for m in filtered
                  if m["date"][5:7] == month and m["type"] == "ingreso")
        exp = sum(m["amount"] for m in filtered
                  if m["date"][5:7] == month and m["type"] != "ingreso")
        cumulative += inc - exp

    # Build rows
    rows = []

    # Saldo anterior special row
    saldo_vals = {m: _fmt(saldo[m]) for m in month_keys}
    saldo_total = sum(saldo_vals.values())
    rows.append({
        "type":    "ingreso",
        "id":      "__saldo_anterior__",
        "label":   "Saldo mes anterior",
        "values":  saldo_vals,
        "total":   saldo_total,
        "avg":     saldo_total / max(1, sum(1 for v in saldo_vals.values() if v != 0)),
        "special": True,
    })

    # Auto income items
    for label in INCOME_AUTO_ITEMS:
        aid = f"__auto_{label}__"
        vals = {m: _fmt(item_month.get(aid, {}).get(m, 0)) for m in month_keys}
        tot = sum(vals.values())
        rows.append({
            "type": "ingreso", "id": aid, "label": label,
            "values": vals, "total": tot,
            "avg": tot / max(1, sum(1 for v in vals.values() if v != 0)),
        })

    # User-configured items
    cat_map = {
        "ingreso":           "ingresos",
        "costo_venta":       "costos_venta",
        "gasto_venta":       "gastos_venta",
        "gasto_operacional": "gastos_operacionales",
        "pago_deuda":        "pago_deudas",
    }

    for typ in TYPE_ORDER:
        cat_key = cat_map[typ]
        for item in project["items"].get(cat_key, []):
            vals = {m: _fmt(item_month.get(item["id"], {}).get(m, 0)) for m in month_keys}
            tot = sum(vals.values())
            rows.append({
                "type": typ, "id": item["id"], "label": item["label"],
                "values": vals, "total": tot,
                "avg": tot / max(1, sum(1 for v in vals.values() if v != 0)),
            })

    # Type subtotals (excluding saldo)
    type_subtotals = {}
    for typ in TYPE_ORDER:
        type_rows = [r for r in rows if r["type"] == typ and not r.get("special")]
        vals = {m: sum(r["values"][m] for r in type_rows) for m in month_keys}
        tot = sum(vals.values())
        type_subtotals[typ] = {
            "values": vals, "total": tot,
            "avg": tot / max(1, sum(1 for v in vals.values() if v != 0)),
        }

    return {"rows": rows, "type_subtotals": type_subtotals, "month_keys": month_keys}


def _apply_styles(ws, row_num: int, col_count: int, fill_color: str,
                  bold=False, font_color="E8ECF4"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=bold, color=font_color)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center" if col > 2 else "left")


def _write_project_sheet(ws, project: dict, movements: list, year: int):
    month_labels = [lbl for _, lbl in MONTHS]
    headers = ["Categoría", "Ítem"] + month_labels + ["Total Anual", "Promedio"]
    col_count = len(headers)

    matrix = build_matrix(project, movements, year)
    rows = matrix["rows"]
    type_subtotals = matrix["type_subtotals"]
    month_keys = matrix["month_keys"]

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        cell.font = Font(bold=True, color="62CBE6")
        cell.alignment = Alignment(horizontal="center" if col > 2 else "left")

    current_row = 2

    def write_row(category_label, item_label, vals, total, avg, fill=None, bold=False):
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=category_label)
        ws.cell(row=current_row, column=2, value=item_label)
        for i, m in enumerate(month_keys):
            ws.cell(row=current_row, column=3 + i, value=vals.get(m, 0))
        ws.cell(row=current_row, column=3 + len(month_keys), value=round(total, 2))
        ws.cell(row=current_row, column=4 + len(month_keys), value=round(avg, 2))
        if fill:
            _apply_styles(ws, current_row, col_count, fill, bold=bold)
        current_row += 1

    # Group by type
    from .defaults import CATEGORY_META
    type_to_label = {v["type"]: v["label"] for v in CATEGORY_META.values()}

    # Saldo anterior
    saldo_row = next((r for r in rows if r["id"] == "__saldo_anterior__"), None)
    if saldo_row:
        write_row("Ingresos", saldo_row["label"], saldo_row["values"],
                  saldo_row["total"], saldo_row["avg"], fill="0F1520")

    for typ in TYPE_ORDER:
        type_rows = [r for r in rows if r["type"] == typ and not r.get("special")]
        for r in type_rows:
            write_row(type_to_label.get(typ, typ), r["label"],
                      r["values"], r["total"], r["avg"])

        # Subtotal
        sub = type_subtotals[typ]
        write_row(f"TOTAL {MOVEMENT_TYPE_LABELS[typ].upper()}", "",
                  sub["values"], sub["total"], sub["avg"],
                  fill=COLORS["total_bg"], bold=True)
        current_row += 1  # blank spacer

    # Total egresos
    exp_vals = {m: sum(type_subtotals[t]["values"].get(m, 0) for t in EXPENSE_TYPES)
                for m in month_keys}
    exp_total = sum(exp_vals.values())
    write_row("TOTAL EGRESOS", "", exp_vals, exp_total, 0,
              fill=COLORS["grand_total_bg"], bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    for i in range(len(month_keys) + 2):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13


def export_project_xlsx(project: dict, movements: list, year: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = project["name"][:31].replace("/", "-")
    _write_project_sheet(ws, project, movements, year)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_all_projects_xlsx(projects: list, all_movements: list, year: int) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for project in projects:
        safe_name = project["name"][:31].replace("/", "-")
        ws = wb.create_sheet(title=safe_name)
        proj_movements = [m for m in all_movements if m["project_id"] == project["id"]]
        _write_project_sheet(ws, project, proj_movements, year)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filtered_xlsx(movements: list, projects: list) -> bytes:
    """Export a flat list of movements as XLSX."""
    project_map = {p["id"]: p["name"] for p in projects}
    from .defaults import MOVEMENT_TYPE_LABELS as MTL

    data = [{
        "Fecha":       m["date"],
        "Proyecto":    project_map.get(m["project_id"], m["project_id"]),
        "Tipo":        MTL.get(m["type"], m["type"]),
        "Ítem":        m["item_label"],
        "Monto":       m["amount"],
        "Descripción": m.get("description", ""),
    } for m in sorted(movements, key=lambda x: x["date"], reverse=True)]

    df = pd.DataFrame(data)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Movimientos")
    return buf.getvalue()


def movements_to_dataframe(movements: list, projects: list) -> pd.DataFrame:
    """Convert movements to a display DataFrame."""
    from .defaults import MOVEMENT_TYPE_LABELS as MTL
    project_map = {p["id"]: p["name"] for p in projects}

    rows = []
    for m in sorted(movements, key=lambda x: x["date"], reverse=True):
        rows.append({
            "Fecha":       m["date"],
            "Proyecto":    project_map.get(m["project_id"], "—"),
            "Tipo":        MTL.get(m["type"], m["type"]),
            "Ítem":        m["item_label"],
            "Monto":       m["amount"],
            "Descripción": m.get("description", ""),
            "_id":         m["id"],
        })
    return pd.DataFrame(rows)
